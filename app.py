#!/usr/bin/env python3
"""
JIRA工单自动化系统 - Python版本
简单高效的JIRA效率分析工具
"""

import os
import json
import pandas as pd
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, send_file
import plotly.graph_objs as go
import plotly.express as px
from plotly.utils import PlotlyJSONEncoder
import requests
from requests.auth import HTTPBasicAuth
import base64

app = Flask(__name__)

# 全局数据存储
jira_data = None
assignees_data = None
jira_connection = None

class JiraConnector:
    """JIRA API连接器"""
    def __init__(self):
        self.server = None
        self.username = None
        self.token = None
        self.session = None
        
    def connect(self, server, username, token):
        """连接到JIRA服务器"""
        try:
            self.server = server.rstrip('/')
            self.username = username
            self.token = token
            
            # 创建会话
            self.session = requests.Session()
            self.session.auth = HTTPBasicAuth(username, token)
            
            # 测试连接
            response = self.session.get(f"{self.server}/rest/api/2/myself")
            if response.status_code == 200:
                return True, "连接成功"
            else:
                return False, f"连接失败: {response.status_code} - {response.text}"
        except Exception as e:
            return False, f"连接错误: {str(e)}"
    
    def get_projects(self):
        """获取所有项目"""
        if not self.session:
            return []
        
        try:
            response = self.session.get(f"{self.server}/rest/api/2/project")
            if response.status_code == 200:
                projects = response.json()
                return [(p['key'], p['name']) for p in projects]
            return []
        except Exception as e:
            print(f"获取项目列表失败: {e}")
            return []
    
    def get_project_issues(self, project_key, max_results=1000):
        """获取项目的所有工单"""
        if not self.session:
            return []
        
        try:
            # JQL查询
            jql = f"project = {project_key} ORDER BY created DESC"
            
            params = {
                'jql': jql,
                'maxResults': max_results,
                'fields': 'key,summary,assignee,priority,status,created,resolutiondate,worklog,timetracking,components,labels,reporter,issuetype'
            }
            
            response = self.session.get(f"{self.server}/rest/api/2/search", params=params)
            
            if response.status_code == 200:
                data = response.json()
                issues = []
                
                for issue in data['issues']:
                    # 获取工单的工作日志
                    worklog_time = self.get_issue_worklog(issue['key'])
                    
                    # 解析工单数据
                    issue_data = {
                        'ticket_id': issue['key'],
                        'jira_key': issue['key'],
                        'summary': issue['fields'].get('summary', ''),
                        'assignee_employee_id': issue['fields']['assignee']['displayName'] if issue['fields'].get('assignee') else 'Unassigned',
                        'assignee_name': issue['fields']['assignee']['displayName'] if issue['fields'].get('assignee') else 'Unassigned',
                        'priority': issue['fields']['priority']['name'] if issue['fields'].get('priority') else 'Medium',
                        'status': issue['fields']['status']['name'],
                        'issue_type': issue['fields']['issuetype']['name'],
                        'reporter': issue['fields']['reporter']['displayName'] if issue['fields'].get('reporter') else '',
                        'created_time': issue['fields']['created'],
                        'resolved_time': issue['fields'].get('resolutiondate'),
                        'log_time': worklog_time,
                        'actual_processing_minutes': worklog_time,
                        'assignment_method': 'MANUAL',  # 默认为手动，可以通过标签或自定义字段判断
                        'components': [c['name'] for c in issue['fields'].get('components', [])],
                        'labels': issue['fields'].get('labels', [])
                    }
                    
                    # 通过标签或组件判断是否AI分单
                    if any('ai' in label.lower() for label in issue_data['labels']) or \
                       any('ai' in comp.lower() for comp in issue_data['components']):
                        issue_data['assignment_method'] = 'AI'
                    
                    issues.append(issue_data)
                
                return issues
            else:
                print(f"获取工单失败: {response.status_code} - {response.text}")
                return []
                
        except Exception as e:
            print(f"获取项目工单失败: {e}")
            return []
    
    def get_issue_worklog(self, issue_key):
        """获取工单的工作日志时间(分钟)"""
        if not self.session:
            return 0
        
        try:
            response = self.session.get(f"{self.server}/rest/api/2/issue/{issue_key}/worklog")
            if response.status_code == 200:
                worklog_data = response.json()
                total_seconds = sum(log.get('timeSpentSeconds', 0) for log in worklog_data.get('worklogs', []))
                return round(total_seconds / 60) if total_seconds > 0 else 0
            return 0
        except Exception as e:
            print(f"获取工作日志失败: {e}")
            return 0

def load_sample_data():
    """加载示例数据"""
    global jira_data, assignees_data
    
    # 示例处理人员数据
    assignees_data = pd.DataFrame([
        {'employee_id': 'EMP001', 'name': '张三', 'department': 'IT支持', 'skill_level': 'SENIOR'},
        {'employee_id': 'EMP002', 'name': '李四', 'department': 'IT支持', 'skill_level': 'INTERMEDIATE'},
        {'employee_id': 'EMP003', 'name': '王五', 'department': '系统运维', 'skill_level': 'EXPERT'},
        {'employee_id': 'EMP004', 'name': '赵六', 'department': '系统运维', 'skill_level': 'JUNIOR'},
        {'employee_id': 'EMP005', 'name': '陈七', 'department': '数据库管理', 'skill_level': 'LEAD'},
    ])
    
    # 示例工单数据
    base_date = datetime.now() - timedelta(days=30)
    sample_tickets = []
    
    for i in range(1, 51):  # 50个示例工单
        days_ago = i % 30
        created_time = base_date + timedelta(days=days_ago)
        assigned_time = created_time + timedelta(hours=2)
        resolved_time = assigned_time + timedelta(hours=4, minutes=(i*10) % 120)
        
        processing_minutes = 50 + (i * 7) % 200
        sample_tickets.append({
            'ticket_id': f'TICKET-{i:03d}',
            'jira_key': f'PROJ-{i:04d}',
            'summary': f'示例工单 {i}',
            'assignee_employee_id': f'EMP{(i % 5) + 1:03d}',
            'priority': ['LOW', 'MEDIUM', 'HIGH', 'CRITICAL'][i % 4],
            'status': ['RESOLVED', 'CLOSED'][i % 2],
            'assignment_method': ['AI', 'MANUAL'][i % 2],
            'created_time': created_time.isoformat(),
            'assigned_time': assigned_time.isoformat(),
            'resolved_time': resolved_time.isoformat(),
            'log_time': processing_minutes,
            'actual_processing_minutes': processing_minutes,
        })
    
    jira_data = pd.DataFrame(sample_tickets)
    jira_data['created_time'] = pd.to_datetime(jira_data['created_time'])
    jira_data['assigned_time'] = pd.to_datetime(jira_data['assigned_time'])
    jira_data['resolved_time'] = pd.to_datetime(jira_data['resolved_time'])

def calculate_efficiency_metrics(data):
    """计算效率指标"""
    if data is None or data.empty:
        return {}
    
    # 基本统计
    total_tickets = len(data)
    ai_tickets = len(data[data['assignment_method'] == 'AI'])
    manual_tickets = len(data[data['assignment_method'] == 'MANUAL'])
    
    # 计算平均处理时间
    avg_ai_time = data[data['assignment_method'] == 'AI']['actual_processing_minutes'].mean()
    avg_manual_time = data[data['assignment_method'] == 'MANUAL']['actual_processing_minutes'].mean()
    
    # 完成率
    completed_tickets = len(data[data['status'].isin(['RESOLVED', 'CLOSED'])])
    completion_rate = (completed_tickets / total_tickets * 100) if total_tickets > 0 else 0
    
    # 增强的效率计算指标
    efficiency_improvement = 0
    time_saved_total = 0
    productivity_gain = 0
    cost_efficiency = 0
    ai_speed_ratio = 0
    
    if not pd.isna(avg_ai_time) and not pd.isna(avg_manual_time) and avg_manual_time > 0:
        # 效率提升百分比
        efficiency_improvement = ((avg_manual_time - avg_ai_time) / avg_manual_time * 100)
        
        # AI处理的工单总计节省的时间（分钟）
        time_saved_total = ai_tickets * (avg_manual_time - avg_ai_time)
        
        # 生产力提升倍数
        productivity_gain = avg_manual_time / avg_ai_time if avg_ai_time > 0 else 1
        
        # AI处理速度相对比率（AI速度是人工的多少倍）
        ai_speed_ratio = avg_manual_time / avg_ai_time if avg_ai_time > 0 else 1
        
        # 成本效益计算（假设每分钟成本为1单位）
        manual_total_cost = manual_tickets * avg_manual_time
        ai_total_cost = ai_tickets * avg_ai_time
        if manual_total_cost > 0:
            cost_efficiency = ((manual_total_cost - ai_total_cost) / manual_total_cost * 100)
    
    return {
        'total_tickets': total_tickets,
        'ai_tickets': ai_tickets,
        'manual_tickets': manual_tickets,
        'avg_ai_processing_time': round(avg_ai_time, 2) if not pd.isna(avg_ai_time) else 0,
        'avg_manual_processing_time': round(avg_manual_time, 2) if not pd.isna(avg_manual_time) else 0,
        'completion_rate': round(completion_rate, 2),
        'efficiency_improvement': round(efficiency_improvement, 2),
        'time_saved_total_minutes': round(time_saved_total, 2),
        'time_saved_total_hours': round(time_saved_total / 60, 2),
        'productivity_gain_ratio': round(productivity_gain, 2),
        'ai_speed_ratio': round(ai_speed_ratio, 2),
        'cost_efficiency_percent': round(cost_efficiency, 2)
    }

@app.route('/')
def dashboard():
    """主页仪表板"""
    global jira_data
    if jira_data is None:
        load_sample_data()
    
    metrics = calculate_efficiency_metrics(jira_data)
    return render_template('dashboard.html', metrics=metrics)

@app.route('/api/metrics')
def api_metrics():
    """API: 获取效率指标"""
    global jira_data
    if jira_data is None:
        load_sample_data()
    
    metrics = calculate_efficiency_metrics(jira_data)
    return jsonify(metrics)

@app.route('/api/charts/comparison')
def api_chart_comparison():
    """API: AI vs 人工分单对比图表"""
    global jira_data
    if jira_data is None:
        load_sample_data()
    
    # 按分单方式分组统计
    comparison = jira_data.groupby('assignment_method').agg({
        'actual_processing_minutes': ['mean', 'count'],
        'ticket_id': 'count'
    }).round(2)
    
    comparison.columns = ['avg_time', 'count', 'total']
    comparison = comparison.reset_index()
    
    # 创建对比图表
    fig = px.bar(
        comparison, 
        x='assignment_method', 
        y='avg_time',
        title='AI vs 人工分单 - 平均处理时间对比',
        labels={'assignment_method': '分单方式', 'avg_time': '平均处理时间(分钟)'}
    )
    
    graphJSON = json.dumps(fig, cls=PlotlyJSONEncoder)
    return jsonify(graphJSON)

@app.route('/api/charts/workload')
def api_chart_workload():
    """API: 工作负载分布图表"""
    global jira_data, assignees_data
    if jira_data is None or assignees_data is None:
        load_sample_data()
    
    # 按处理人员统计工单数量
    workload = jira_data['assignee_employee_id'].value_counts().reset_index()
    workload.columns = ['employee_id', 'ticket_count']
    
    # 合并处理人员姓名
    workload = workload.merge(assignees_data[['employee_id', 'name']], 
                             left_on='employee_id', right_on='employee_id', how='left')
    
    # 创建工作负载图表
    fig = px.pie(
        workload, 
        values='ticket_count', 
        names='name',
        title='处理人员工作负载分布'
    )
    
    graphJSON = json.dumps(fig, cls=PlotlyJSONEncoder)
    return jsonify(graphJSON)

@app.route('/api/upload', methods=['POST'])
def api_upload():
    """API: 上传JIRA数据文件"""
    global jira_data
    
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    try:
        # 根据文件类型读取数据，处理编码问题
        if file.filename.endswith('.csv'):
            # 保存原始文件内容
            file.seek(0)
            file_content = file.read()
            
            # 尝试多种编码方式读取CSV文件
            encodings_to_try = [
                'utf-8', 'utf-8-sig', 
                'gbk', 'gb2312', 'gb18030', 
                'big5', 'ascii',
                'iso-8859-1', 'cp1252', 'latin1',
                'windows-1252', 'ansi'
            ]
            
            jira_data = None
            last_error = None
            
            for encoding in encodings_to_try:
                try:
                    # 使用字节内容创建StringIO对象
                    import io
                    decoded_content = file_content.decode(encoding)
                    csv_file = io.StringIO(decoded_content)
                    jira_data = pd.read_csv(csv_file)
                    print(f"成功使用 {encoding} 编码读取文件")
                    break
                except (UnicodeDecodeError, UnicodeError) as e:
                    last_error = f"{encoding}: {str(e)}"
                    continue
                except Exception as e:
                    last_error = f"{encoding}: {str(e)}"
                    # 如果不是编码错误，继续尝试其他编码
                    continue
            
            if jira_data is None:
                # 尝试使用chardet自动检测编码
                try:
                    import chardet
                    detected = chardet.detect(file_content)
                    detected_encoding = detected['encoding']
                    confidence = detected['confidence']
                    
                    if detected_encoding and confidence > 0.7:
                        print(f"检测到编码: {detected_encoding} (置信度: {confidence:.2f})")
                        try:
                            import io
                            decoded_content = file_content.decode(detected_encoding)
                            csv_file = io.StringIO(decoded_content)
                            jira_data = pd.read_csv(csv_file)
                            print(f"成功使用检测到的编码 {detected_encoding} 读取文件")
                        except Exception as e:
                            print(f"使用检测到的编码失败: {e}")
                except ImportError:
                    print("chardet库未安装，跳过自动编码检测")
                except Exception as e:
                    print(f"自动编码检测失败: {e}")
            
            if jira_data is None:
                # 最后尝试：使用多种方法忽略错误字符
                fallback_methods = [
                    ('utf-8', 'ignore'),
                    ('gbk', 'ignore'), 
                    ('gb18030', 'ignore'),
                    ('iso-8859-1', 'ignore'),
                    ('cp1252', 'replace')
                ]
                
                for encoding, error_handling in fallback_methods:
                    try:
                        import io
                        decoded_content = file_content.decode(encoding, errors=error_handling)
                        csv_file = io.StringIO(decoded_content)
                        jira_data = pd.read_csv(csv_file)
                        print(f"使用 {encoding} ({error_handling} 错误字符) 成功读取文件")
                        break
                    except Exception as e:
                        continue
            
            if jira_data is None:
                return jsonify({
                    'error': f'无法读取CSV文件，已尝试多种编码方式。建议: 1)用Excel打开文件，另存为UTF-8编码的CSV；2)检查文件是否损坏；3)确认文件确实是CSV格式。详细错误: {last_error}'
                }), 400
                
        elif file.filename.endswith('.xlsx'):
            jira_data = pd.read_excel(file)
        elif file.filename.endswith('.json'):
            data = json.load(file)
            if 'tickets' in data:
                jira_data = pd.DataFrame(data['tickets'])
            else:
                jira_data = pd.DataFrame(data)
        else:
            return jsonify({'error': '不支持的文件格式'}), 400
        
        # 数据预处理
        if 'created_time' in jira_data.columns:
            jira_data['created_time'] = pd.to_datetime(jira_data['created_time'])
        if 'assigned_time' in jira_data.columns:
            jira_data['assigned_time'] = pd.to_datetime(jira_data['assigned_time'])
        if 'resolved_time' in jira_data.columns:
            jira_data['resolved_time'] = pd.to_datetime(jira_data['resolved_time'])
        
        # 智能处理时间字段：优先使用log_time，如果没有则使用actual_processing_minutes
        if 'log_time' in jira_data.columns and 'actual_processing_minutes' not in jira_data.columns:
            jira_data['actual_processing_minutes'] = jira_data['log_time']
        elif 'log_time' in jira_data.columns and 'actual_processing_minutes' in jira_data.columns:
            # 如果两个字段都存在，优先使用log_time（如果不为空）
            jira_data['actual_processing_minutes'] = jira_data['log_time'].fillna(jira_data['actual_processing_minutes'])
        
        return jsonify({
            'success': True,
            'message': f'成功上传 {len(jira_data)} 条记录',
            'rows': len(jira_data),
            'columns': list(jira_data.columns)
        })
        
    except Exception as e:
        return jsonify({'error': f'文件处理失败: {str(e)}'}), 500

@app.route('/api/efficiency/analysis')
def api_efficiency_analysis():
    """API: 详细的效率分析对比"""
    global jira_data
    if jira_data is None:
        load_sample_data()
    
    metrics = calculate_efficiency_metrics(jira_data)
    
    # 构建详细的效率分析报告
    analysis_report = {
        'summary': {
            'total_tickets': metrics['total_tickets'],
            'ai_vs_manual_ratio': f"{metrics['ai_tickets']}:{metrics['manual_tickets']}",
            'overall_efficiency_improvement': metrics['efficiency_improvement']
        },
        'time_analysis': {
            'ai_avg_time': metrics['avg_ai_processing_time'],
            'manual_avg_time': metrics['avg_manual_processing_time'],
            'time_saved_per_ticket': round(metrics['avg_manual_processing_time'] - metrics['avg_ai_processing_time'], 2),
            'total_time_saved_hours': metrics['time_saved_total_hours'],
            'ai_speed_advantage': f"{metrics['ai_speed_ratio']}x faster"
        },
        'productivity_metrics': {
            'productivity_gain_ratio': metrics['productivity_gain_ratio'],
            'cost_efficiency': metrics['cost_efficiency_percent'],
            'completion_rate': metrics['completion_rate']
        },
        'roi_calculation': {
            'description': "基于处理时间节省计算的投资回报",
            'formula': "ROI = (节省时间成本 - AI系统成本) / AI系统成本 × 100%",
            'estimated_monthly_savings': round(metrics['time_saved_total_hours'] * 50, 2),  # 假设每小时成本50元
            'efficiency_score': min(100, max(0, metrics['efficiency_improvement']))
        }
    }
    
    return jsonify(analysis_report)

@app.route('/api/jira/connect', methods=['POST'])
def api_jira_connect():
    """API: 连接JIRA服务器"""
    global jira_connection
    
    data = request.get_json()
    if not data:
        return jsonify({'error': '缺少连接参数'}), 400
    
    server = data.get('server')
    username = data.get('username') 
    token = data.get('token')
    
    if not all([server, username, token]):
        return jsonify({'error': '请提供完整的连接参数: server, username, token'}), 400
    
    try:
        jira_connection = JiraConnector()
        success, message = jira_connection.connect(server, username, token)
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'server': server,
                'username': username
            })
        else:
            return jsonify({'error': message}), 400
            
    except Exception as e:
        return jsonify({'error': f'连接失败: {str(e)}'}), 500

@app.route('/api/jira/projects')
def api_jira_projects():
    """API: 获取JIRA项目列表"""
    global jira_connection
    
    if not jira_connection:
        return jsonify({'error': '请先连接JIRA服务器'}), 400
    
    try:
        projects = jira_connection.get_projects()
        return jsonify({
            'success': True,
            'projects': projects,
            'count': len(projects)
        })
    except Exception as e:
        return jsonify({'error': f'获取项目列表失败: {str(e)}'}), 500

@app.route('/api/jira/import/<project_key>')
def api_jira_import(project_key):
    """API: 从JIRA导入项目数据"""
    global jira_connection, jira_data
    
    if not jira_connection:
        return jsonify({'error': '请先连接JIRA服务器'}), 400
    
    max_results = request.args.get('max_results', 1000, type=int)
    
    try:
        issues = jira_connection.get_project_issues(project_key, max_results)
        
        if not issues:
            return jsonify({'error': '未找到工单数据或项目不存在'}), 404
        
        # 转换为DataFrame
        jira_data = pd.DataFrame(issues)
        
        # 数据预处理
        jira_data['created_time'] = pd.to_datetime(jira_data['created_time'])
        if 'resolved_time' in jira_data.columns:
            jira_data['resolved_time'] = pd.to_datetime(jira_data['resolved_time'])
        
        return jsonify({
            'success': True,
            'message': f'成功导入 {len(issues)} 条工单数据',
            'project_key': project_key,
            'total_issues': len(issues),
            'issues_with_worklog': len([i for i in issues if i['log_time'] > 0]),
            'issue_types': list(set(i['issue_type'] for i in issues)),
            'assignees': list(set(i['assignee_name'] for i in issues if i['assignee_name'] != 'Unassigned'))
        })
        
    except Exception as e:
        return jsonify({'error': f'导入数据失败: {str(e)}'}), 500

@app.route('/api/jira/analysis/advanced')
def api_jira_advanced_analysis():
    """API: JIRA项目管理专业分析"""
    global jira_data
    if jira_data is None or jira_data.empty:
        return jsonify({'error': '没有数据，请先导入JIRA项目数据'}), 400
    
    try:
        analysis = {
            # 1. 项目健康度分析
            'project_health': analyze_project_health(jira_data),
            
            # 2. 团队绩效分析
            'team_performance': analyze_team_performance(jira_data),
            
            # 3. 工单流转分析
            'workflow_analysis': analyze_workflow(jira_data),
            
            # 4. 质量指标分析
            'quality_metrics': analyze_quality_metrics(jira_data),
            
            # 5. 资源分配分析
            'resource_allocation': analyze_resource_allocation(jira_data),
            
            # 6. 趋势预测分析
            'trend_prediction': analyze_trends(jira_data),
            
            # 7. 关键洞察和建议
            'insights_and_recommendations': generate_insights(jira_data)
        }
        
        return jsonify(analysis)
        
    except Exception as e:
        return jsonify({'error': f'分析失败: {str(e)}'}), 500

def analyze_project_health(data):
    """项目健康度分析"""
    total_issues = len(data)
    resolved_issues = len(data[data['status'].isin(['Resolved', 'Closed', 'Done'])])
    in_progress_issues = len(data[data['status'].isin(['In Progress', 'In Review'])])
    
    # 计算健康度得分
    resolution_rate = (resolved_issues / total_issues * 100) if total_issues > 0 else 0
    
    # 平均解决时间
    resolved_data = data[data['resolved_time'].notna()]
    if not resolved_data.empty:
        avg_resolution_days = (resolved_data['resolved_time'] - resolved_data['created_time']).dt.days.mean()
    else:
        avg_resolution_days = 0
    
    health_score = min(100, max(0, resolution_rate - (avg_resolution_days * 2)))
    
    return {
        'total_issues': total_issues,
        'resolved_issues': resolved_issues,
        'in_progress_issues': in_progress_issues,
        'resolution_rate': round(resolution_rate, 2),
        'avg_resolution_days': round(avg_resolution_days, 1),
        'health_score': round(health_score, 1),
        'health_status': get_health_status(health_score)
    }

def analyze_team_performance(data):
    """团队绩效分析"""
    if 'assignee_name' not in data.columns:
        return {}
    
    team_stats = data.groupby('assignee_name').agg({
        'ticket_id': 'count',
        'log_time': ['mean', 'sum'],
        'priority': lambda x: (x == 'High').sum(),
        'status': lambda x: x.isin(['Resolved', 'Closed', 'Done']).sum()
    }).round(2)
    
    team_stats.columns = ['total_tickets', 'avg_time', 'total_time', 'high_priority_count', 'resolved_count']
    team_stats['resolution_rate'] = (team_stats['resolved_count'] / team_stats['total_tickets'] * 100).round(2)
    team_stats['efficiency_score'] = ((team_stats['resolved_count'] / team_stats['avg_time']) * 100).round(2)
    
    return {
        'team_statistics': team_stats.to_dict('index'),
        'top_performer': team_stats['efficiency_score'].idxmax(),
        'most_productive': team_stats['total_tickets'].idxmax(),
        'team_size': len(team_stats)
    }

def analyze_workflow(data):
    """工单流转分析"""
    status_distribution = data['status'].value_counts().to_dict()
    priority_distribution = data['priority'].value_counts().to_dict()
    
    # 工单类型分析
    if 'issue_type' in data.columns:
        type_distribution = data['issue_type'].value_counts().to_dict()
    else:
        type_distribution = {}
    
    # 平均处理时间按优先级
    priority_avg_time = data.groupby('priority')['log_time'].mean().round(2).to_dict()
    
    return {
        'status_distribution': status_distribution,
        'priority_distribution': priority_distribution,
        'type_distribution': type_distribution,
        'avg_time_by_priority': priority_avg_time,
        'bottlenecks': identify_bottlenecks(data)
    }

def analyze_quality_metrics(data):
    """质量指标分析"""
    # 重新打开的工单数量（质量问题指标）
    reopened_count = 0  # 需要通过历史记录或状态变更来统计
    
    # 高优先级工单比例
    high_priority_rate = (data['priority'] == 'High').sum() / len(data) * 100
    
    # 未分配工单比例  
    unassigned_rate = (data['assignee_name'] == 'Unassigned').sum() / len(data) * 100
    
    # 超期工单分析（假设7天为标准处理时间）
    if 'resolved_time' in data.columns:
        resolved_data = data[data['resolved_time'].notna()]
        if not resolved_data.empty:
            processing_days = (resolved_data['resolved_time'] - resolved_data['created_time']).dt.days
            overdue_count = (processing_days > 7).sum()
            overdue_rate = overdue_count / len(resolved_data) * 100
        else:
            overdue_rate = 0
    else:
        overdue_rate = 0
    
    return {
        'high_priority_rate': round(high_priority_rate, 2),
        'unassigned_rate': round(unassigned_rate, 2),
        'overdue_rate': round(overdue_rate, 2),
        'quality_score': round(100 - high_priority_rate - unassigned_rate - overdue_rate, 2)
    }

def analyze_resource_allocation(data):
    """资源分配分析"""
    if 'assignee_name' not in data.columns:
        return {}
    
    # 工作负载分布
    workload = data['assignee_name'].value_counts()
    workload_std = workload.std()
    workload_balance = 100 - min(100, workload_std / workload.mean() * 50)
    
    # 技能匹配分析（基于处理时间）
    skill_analysis = data.groupby('assignee_name')['log_time'].agg(['mean', 'std']).round(2)
    
    return {
        'workload_distribution': workload.to_dict(),
        'workload_balance_score': round(workload_balance, 2),
        'skill_analysis': skill_analysis.to_dict('index'),
        'resource_utilization': calculate_resource_utilization(data)
    }

def analyze_trends(data):
    """趋势分析"""
    if 'created_time' not in data.columns:
        return {}
    
    # 按月统计工单创建趋势
    data['month'] = data['created_time'].dt.to_period('M')
    monthly_creation = data.groupby('month').size().to_dict()
    
    # 解决趋势
    if 'resolved_time' in data.columns:
        resolved_data = data[data['resolved_time'].notna()]
        resolved_data['resolved_month'] = resolved_data['resolved_time'].dt.to_period('M')
        monthly_resolution = resolved_data.groupby('resolved_month').size().to_dict()
    else:
        monthly_resolution = {}
    
    return {
        'monthly_creation_trend': {str(k): v for k, v in monthly_creation.items()},
        'monthly_resolution_trend': {str(k): v for k, v in monthly_resolution.items()},
        'trend_direction': calculate_trend_direction(monthly_creation)
    }

def generate_insights(data):
    """生成关键洞察和建议"""
    insights = []
    
    # 效率洞察
    if data['log_time'].mean() > 0:
        avg_time = data['log_time'].mean()
        if avg_time > 120:  # 超过2小时
            insights.append({
                'type': 'efficiency',
                'level': 'warning',
                'message': f'平均处理时间较长({avg_time:.1f}分钟)，建议优化工作流程',
                'recommendation': '考虑工单分类优化、技能培训或工具改进'
            })
    
    # 工作负载洞察
    if 'assignee_name' in data.columns:
        workload = data['assignee_name'].value_counts()
        if workload.std() / workload.mean() > 0.5:
            insights.append({
                'type': 'workload',
                'level': 'warning', 
                'message': '团队工作负载分配不均衡',
                'recommendation': '考虑重新分配工作或增加资源'
            })
    
    # 质量洞察
    high_priority_rate = (data['priority'] == 'High').sum() / len(data) * 100
    if high_priority_rate > 30:
        insights.append({
            'type': 'quality',
            'level': 'alert',
            'message': f'高优先级工单比例过高({high_priority_rate:.1f}%)',
            'recommendation': '需要加强需求管理和问题预防'
        })
    
    return insights

def get_health_status(score):
    """获取健康状态"""
    if score >= 80:
        return '优秀'
    elif score >= 60:
        return '良好'
    elif score >= 40:
        return '一般'
    else:
        return '需要改进'

def identify_bottlenecks(data):
    """识别瓶颈"""
    bottlenecks = []
    
    # 状态瓶颈
    in_progress = data[data['status'].isin(['In Progress', 'In Review'])]
    if len(in_progress) > len(data) * 0.3:
        bottlenecks.append('处理中工单积压严重')
    
    return bottlenecks

def calculate_resource_utilization(data):
    """计算资源利用率"""
    if 'assignee_name' not in data.columns:
        return 0
    
    assigned = len(data[data['assignee_name'] != 'Unassigned'])
    total = len(data)
    return round(assigned / total * 100, 2) if total > 0 else 0

def calculate_trend_direction(monthly_data):
    """计算趋势方向"""
    if len(monthly_data) < 2:
        return 'stable'
    
    values = list(monthly_data.values())
    recent_avg = sum(values[-3:]) / len(values[-3:]) if len(values) >= 3 else values[-1]
    earlier_avg = sum(values[:-3]) / len(values[:-3]) if len(values) > 3 else values[0]
    
    if recent_avg > earlier_avg * 1.1:
        return 'increasing'
    elif recent_avg < earlier_avg * 0.9:
        return 'decreasing' 
    else:
        return 'stable'

@app.route('/api/template/csv')
def api_download_csv_template():
    """API: 下载CSV数据模板"""
    import io
    import csv
    from flask import make_response
    
    # 创建CSV模板数据
    template_data = [
        ['ticket_id', 'jira_key', 'summary', 'assignee_employee_id', 'priority', 'status', 'assignment_method', 'created_time', 'assigned_time', 'resolved_time', 'log_time', 'actual_processing_minutes'],
        ['TICKET-001', 'PROJ-1001', '系统登录问题', 'EMP001', 'HIGH', 'RESOLVED', 'AI', '2024-01-15 10:00:00', '2024-01-15 10:05:00', '2024-01-15 11:00:00', '55', '55'],
        ['TICKET-002', 'PROJ-1002', '数据库连接超时', 'EMP002', 'CRITICAL', 'RESOLVED', 'MANUAL', '2024-01-15 11:00:00', '2024-01-15 11:30:00', '2024-01-15 13:00:00', '90', '90'],
        ['TICKET-003', 'PROJ-1003', '界面显示异常', 'EMP001', 'MEDIUM', 'RESOLVED', 'AI', '2024-01-15 14:00:00', '2024-01-15 14:02:00', '2024-01-15 14:45:00', '43', '43'],
        ['TICKET-004', 'PROJ-1004', '报表导出失败', 'EMP003', 'LOW', 'RESOLVED', 'MANUAL', '2024-01-15 15:00:00', '2024-01-15 16:00:00', '2024-01-15 17:30:00', '90', '90'],
        ['TICKET-005', 'PROJ-1005', '邮件发送失败', 'EMP002', 'HIGH', 'RESOLVED', 'AI', '2024-01-16 09:00:00', '2024-01-16 09:03:00', '2024-01-16 09:30:00', '27', '27']
    ]
    
    # 创建CSV字符串
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(template_data)
    csv_content = output.getvalue()
    output.close()
    
    # 创建响应
    response = make_response(csv_content)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename="jira_data_template.csv"'
    
    return response

@app.route('/api/template/excel')
def api_download_excel_template():
    """API: 下载Excel数据模板"""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # 创建工作簿
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 示例数据
        template_data = pd.DataFrame([
            {
                'ticket_id': 'TICKET-001',
                'jira_key': 'PROJ-1001', 
                'summary': '系统登录问题',
                'assignee_employee_id': 'EMP001',
                'priority': 'HIGH',
                'status': 'RESOLVED',
                'assignment_method': 'AI',
                'created_time': '2024-01-15 10:00:00',
                'assigned_time': '2024-01-15 10:05:00', 
                'resolved_time': '2024-01-15 11:00:00',
                'log_time': 55,
                'actual_processing_minutes': 55
            },
            {
                'ticket_id': 'TICKET-002',
                'jira_key': 'PROJ-1002',
                'summary': '数据库连接超时', 
                'assignee_employee_id': 'EMP002',
                'priority': 'CRITICAL',
                'status': 'RESOLVED',
                'assignment_method': 'MANUAL',
                'created_time': '2024-01-15 11:00:00',
                'assigned_time': '2024-01-15 11:30:00',
                'resolved_time': '2024-01-15 13:00:00',
                'log_time': 90, 
                'actual_processing_minutes': 90
            },
            {
                'ticket_id': 'TICKET-003',
                'jira_key': 'PROJ-1003',
                'summary': '界面显示异常',
                'assignee_employee_id': 'EMP001', 
                'priority': 'MEDIUM',
                'status': 'RESOLVED',
                'assignment_method': 'AI',
                'created_time': '2024-01-15 14:00:00',
                'assigned_time': '2024-01-15 14:02:00',
                'resolved_time': '2024-01-15 14:45:00',
                'log_time': 43,
                'actual_processing_minutes': 43
            },
            {
                'ticket_id': 'TICKET-004', 
                'jira_key': 'PROJ-1004',
                'summary': '报表导出失败',
                'assignee_employee_id': 'EMP003',
                'priority': 'LOW', 
                'status': 'RESOLVED',
                'assignment_method': 'MANUAL',
                'created_time': '2024-01-15 15:00:00',
                'assigned_time': '2024-01-15 16:00:00',
                'resolved_time': '2024-01-15 17:30:00',
                'log_time': 90,
                'actual_processing_minutes': 90
            },
            {
                'ticket_id': 'TICKET-005',
                'jira_key': 'PROJ-1005', 
                'summary': '邮件发送失败',
                'assignee_employee_id': 'EMP002',
                'priority': 'HIGH',
                'status': 'RESOLVED', 
                'assignment_method': 'AI',
                'created_time': '2024-01-16 09:00:00',
                'assigned_time': '2024-01-16 09:03:00',
                'resolved_time': '2024-01-16 09:30:00',
                'log_time': 27,
                'actual_processing_minutes': 27
            }
        ])
        
        # 创建数据模板工作表
        ws1 = wb.create_sheet("数据模板")
        for r in dataframe_to_rows(template_data, index=False, header=True):
            ws1.append(r)
        
        # 创建字段说明工作表
        field_descriptions = pd.DataFrame([
            {'字段名': 'ticket_id', '说明': '工单唯一标识符', '示例': 'TICKET-001', '必填': '是'},
            {'字段名': 'jira_key', '说明': 'JIRA系统中的工单键', '示例': 'PROJ-1001', '必填': '否'},
            {'字段名': 'summary', '说明': '工单标题/摘要', '示例': '系统登录问题', '必填': '否'},
            {'字段名': 'assignee_employee_id', '说明': '处理人员ID', '示例': 'EMP001', '必填': '是'},
            {'字段名': 'priority', '说明': '优先级', '示例': 'LOW/MEDIUM/HIGH/CRITICAL', '必填': '否'},
            {'字段名': 'status', '说明': '工单状态', '示例': 'OPEN/RESOLVED/CLOSED', '必填': '否'}, 
            {'字段名': 'assignment_method', '说明': '分单方式', '示例': 'AI/MANUAL', '必填': '是'},
            {'字段名': 'created_time', '说明': '创建时间', '示例': '2024-01-15 10:00:00', '必填': '否'},
            {'字段名': 'assigned_time', '说明': '分配时间', '示例': '2024-01-15 10:05:00', '必填': '否'},
            {'字段名': 'resolved_time', '说明': '解决时间', '示例': '2024-01-15 11:00:00', '必填': '否'},
            {'字段名': 'log_time', '说明': '记录的完成时间(分钟)', '示例': '55', '必填': '是'},
            {'字段名': 'actual_processing_minutes', '说明': '实际处理时间(分钟)', '示例': '55', '必填': '是'}
        ])
        
        ws2 = wb.create_sheet("字段说明")
        for r in dataframe_to_rows(field_descriptions, index=False, header=True):
            ws2.append(r)
        
        # 保存到内存中的字节流
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 创建响应
        from flask import Response
        response = Response(
            excel_buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=jira_data_template.xlsx'}
        )
        
        return response
        
    except Exception as e:
        return jsonify({'error': f'生成Excel模板失败: {str(e)}'}), 500

@app.route('/api/export/excel')
def api_export_excel():
    """API: 导出分析报告为Excel"""
    global jira_data
    if jira_data is None:
        load_sample_data()
    
    # 创建Excel文件
    output_file = 'jira_analysis_report.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 原始数据
        jira_data.to_excel(writer, sheet_name='原始数据', index=False)
        
        # 效率指标
        metrics = calculate_efficiency_metrics(jira_data)
        metrics_df = pd.DataFrame([metrics])
        metrics_df.to_excel(writer, sheet_name='效率指标', index=False)
        
        # 按分单方式统计
        method_stats = jira_data.groupby('assignment_method').agg({
            'actual_processing_minutes': ['mean', 'std', 'count'],
            'ticket_id': 'count'
        }).round(2)
        method_stats.to_excel(writer, sheet_name='分单方式统计')
        
        # 按处理人员统计
        assignee_stats = jira_data.groupby('assignee_employee_id').agg({
            'actual_processing_minutes': ['mean', 'count'],
            'ticket_id': 'count'
        }).round(2)
        assignee_stats.to_excel(writer, sheet_name='处理人员统计')
    
    return send_file(output_file, as_attachment=True)

if __name__ == '__main__':
    # 确保模板目录存在
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static', exist_ok=True)
    
    print("JIRA Efficiency Analysis System Starting...")
    print("Web Interface: http://localhost:5000")
    print("API Documentation: http://localhost:5000/api/metrics")
    
    app.run(debug=True, host='0.0.0.0', port=5000)